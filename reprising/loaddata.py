import pandas as pd
import os
import numpy as np
from openpyxl import load_workbook
from datetime import datetime


def load_forreprice_data(engine):
    # Функция для приведения значений к одному регистру и удаления пробелов
    def normalize_kod_column(df, column="kod"):
        df[column] = df[column].str.strip().str.upper()
        return df

    # Загрузка и нормализация данных из таблиц
    def load_and_normalize_table(query, column="kod"):
        df = pd.read_sql(query, engine)
        df = normalize_kod_column(df, column)
        return df

    # Загрузка данных с нормализацией
    nomenklatura_query = """
    SELECT kod, artikul, proizvoditel, naimenovaniepolnoe, 
           bazovajaedinitsa AS edizm, datasozdanija
    FROM nomenklatura
    WHERE vidnomenklatury = 'Товар'
    """
    nomenklatura_df = load_and_normalize_table(nomenklatura_query)

    nomenklaturaold_query = """
    SELECT kod, stellazh, pometkaudalenija
    FROM nomenklaturaold
    """
    nomenklaturaold_df = load_and_normalize_table(nomenklaturaold_query)

    deliveryminprice_query = """
    SELECT kod, price AS delprice, sklad AS delsklad
    FROM deliveryminprice
    """
    deliveryminprice_df = load_and_normalize_table(deliveryminprice_query)

    middlemaxprice_query = """
    SELECT kod, middleprice, maxprice
    FROM middlemaxprice
    """
    middlemaxprice_df = load_and_normalize_table(middlemaxprice_query)

    priceold_query = """
    SELECT kod, tsenazakup, tsenarozn
    FROM priceold
    """
    priceold_df = load_and_normalize_table(priceold_query)

    stockold_query = """
    SELECT kod, osnsklad AS ostatok
    FROM stockold
    """
    stockold_df = load_and_normalize_table(stockold_query)

    typedetailgen_query = """
    SELECT kod, type_detail
    FROM typedetailgen
    """
    typedetailgen_df = load_and_normalize_table(typedetailgen_query)

    groupanalogiold_query = """
    SELECT kod_1s AS kod, gruppa_analogov
    FROM groupanalogiold
    """
    groupanalogiold_df = load_and_normalize_table(groupanalogiold_query)

    # Соединение таблиц по колонке 'kod' с использованием левого соединения
    merged_df = (
        nomenklatura_df.merge(nomenklaturaold_df, on="kod", how="left")
        .merge(deliveryminprice_df, on="kod", how="left")
        .merge(middlemaxprice_df, on="kod", how="left")
        .merge(priceold_df, on="kod", how="left")
        .merge(stockold_df, on="kod", how="left")
        .merge(typedetailgen_df, on="kod", how="left")
        .merge(groupanalogiold_df, on="kod", how="left")
    )

    # Заменяем текстовые 'NaN', пустые строки и 'None' на np.nan
    merged_df["gruppa_analogov"] = merged_df["gruppa_analogov"].replace(
        ["NaN", "None", ""], np.nan
    )

    # Присвоение уникальных кодов, если gruppa_analogov имеет NaN
    def assign_unique_codes(df, column):
        unique_code_start = 1000000  # Начало уникальных кодов
        df_copy = df.copy()  # Копируем DataFrame перед изменениями
        df_copy[column] = df_copy[column].fillna(
            pd.Series(
                [unique_code_start + i for i in range(df_copy[column].isna().sum())],
                index=df_copy[df_copy[column].isna()].index,
            )
        )
        return df_copy

    merged_df = assign_unique_codes(merged_df, "gruppa_analogov")

    # Переименовываем колонку naimenovaniepolnoe в naimenovanie
    merged_df.rename(columns={"naimenovaniepolnoe": "naimenovanie"}, inplace=True)

    # Фильтрация по остаткам (ostatok > 0)
    filtered_df = merged_df[merged_df["ostatok"] > 0]

    # Сортировка по алфавиту в колонке naimenovanie
    filtered_df = filtered_df.sort_values(by="naimenovanie")

    # Получение списка kod из filtered_df
    filtered_kods = filtered_df["kod"].tolist()

    # Загрузка данных из других таблиц с учетом нормализации и фильтрации
    priceendmonth_query = """
    SELECT kod, data AS year_month, tsena
    FROM priceendmonth
    """
    priceendmonth_df = load_and_normalize_table(priceendmonth_query)

    # Загрузка текущего месяца и года
    current_month = datetime.now().strftime("%Y-%m")

    # Если для текущего месяца цена равна 0, заменяем её на tsenarozn из filtered_df
    priceendmonth_df = priceendmonth_df.merge(
        filtered_df[["kod", "tsenarozn"]], on="kod", how="left"
    )
    priceendmonth_df.loc[
        (priceendmonth_df["year_month"] == current_month)
        & (priceendmonth_df["tsena"] == 0),
        "tsena",
    ] = priceendmonth_df["tsenarozn"]
    priceendmonth_df.drop(columns=["tsenarozn"], inplace=True)

    salespivot_query = """
    SELECT kod, year_month, kolichestvo, summa
    FROM salespivot
    """
    salespivot_df = load_and_normalize_table(salespivot_query)

    stockendmonth_query = """
    SELECT nomenklaturakod AS kod, month AS year_month, balance AS stock
    FROM stockendmonth
    """
    stockendmonth_df = load_and_normalize_table(stockendmonth_query)

    suppliespivot_query = """
    SELECT kod, year_month, kolichestvo, summa
    FROM suppliespivot
    """
    suppliespivot_df = load_and_normalize_table(suppliespivot_query)

    postuplenija_query = """
    SELECT kod, data, kolichestvo
    FROM postuplenija
    WHERE proveden = 'Да' AND hozoperatsija = 'Поступление товаров'
    """
    postuplenija_df = load_and_normalize_table(postuplenija_query)

    # Фильтрация данных по kod, которые есть в filtered_df
    priceendmonth_filtered = priceendmonth_df[
        priceendmonth_df["kod"].isin(filtered_kods)
    ]
    salespivot_filtered = salespivot_df[salespivot_df["kod"].isin(filtered_kods)]
    stockendmonth_filtered = stockendmonth_df[
        stockendmonth_df["kod"].isin(filtered_kods)
    ]
    suppliespivot_filtered = suppliespivot_df[
        suppliespivot_df["kod"].isin(filtered_kods)
    ]
    postuplenija_filtered = postuplenija_df[postuplenija_df["kod"].isin(filtered_kods)]

    # Возвращаем все датафреймы
    return (
        filtered_df,
        priceendmonth_filtered,
        salespivot_filtered,
        stockendmonth_filtered,
        suppliespivot_filtered,
        postuplenija_filtered,
    )


def exclude_kods_from_file(filtered_df):
    """
    Функция для исключения позиций на основе файла 'Непереоценивать.xlsx', используя лист 'Код'.
    Если файл доступен, проводит объединение с filtered_df и исключает позиции на основе найденных kod.
    При этом сохраняются все листы и данные, включая 'Бренды' и 'Текст'.

    Parameters:
    filtered_df (pd.DataFrame): Основной датафрейм с информацией о позициях.

    Returns:
    pd.DataFrame: Обновленный датафрейм, в котором исключены позиции на основе файла.
    """
    # Пути к файлу
    paths = [
        r"D:\NAS\заказы\Непереоценивать.xlsx",
        r"\\26.218.196.12\заказы\Непереоценивать.xlsx",
    ]

    # Попытка загрузить файл по одному из путей
    file_path = None
    for path in paths:
        if os.path.exists(path):
            file_path = path
            break

    # Если файл не найден ни по одному пути, пропускаем выполнение функции
    if file_path is None:
        print(
            "Файл 'Непереоценивать.xlsx' не найден по указанным путям. Пропуск выполнения функции."
        )
        return filtered_df

    try:
        # Загружаем данные с листа "Код", читая 'kod' как текст
        df_to_exclude = pd.read_excel(file_path, sheet_name="Код", dtype={"kod": str})

        # Приводим колонку 'kod' к строковому типу и удаляем пробелы
        df_to_exclude["kod"] = df_to_exclude["kod"].astype(str).str.strip()

        # Делаем левое соединение с filtered_df по колонке 'kod'
        merged_df = pd.merge(
            filtered_df, df_to_exclude[["kod"]], on="kod", how="left", indicator=True
        )

        # Определяем коды, которые найдены и отсутствуют
        found_kods = merged_df[merged_df["_merge"] == "both"]
        not_found_kods = merged_df[merged_df["_merge"] == "right_only"]

        # Если есть коды, которые отсутствуют в filtered_df, выгружаем их обратно
        if not not_found_kods.empty:
            print(f"Коды, не найденные в filtered_df: {len(not_found_kods)}")
            not_found_df = df_to_exclude[
                df_to_exclude["kod"].isin(not_found_kods["kod"])
            ]

            # Загружаем все листы в словарь для сохранения их данных
            with pd.ExcelWriter(
                file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay"
            ) as writer:
                workbook = load_workbook(file_path)
                workbook.remove(workbook["Код"])  # Удаляем лист "Код" для перезаписи
                workbook.save(file_path)
                not_found_df.to_excel(writer, sheet_name="Код", index=False)
            return filtered_df

        # Оставляем только те строки, где kod из df_to_exclude присутствует в filtered_df
        filtered_exclude = merged_df[merged_df["_merge"] == "both"]

        # Выбираем нужные колонки для экспорта
        result_df = filtered_exclude[
            ["kod", "artikul", "proizvoditel", "naimenovanie", "edizm", "datasozdanija"]
        ]

        # Загружаем все листы из файла
        with pd.ExcelWriter(
            file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay"
        ) as writer:
            workbook = load_workbook(file_path)
            workbook.remove(workbook["Код"])  # Удаляем лист "Код" для перезаписи
            workbook.save(file_path)
            result_df.to_excel(
                writer, sheet_name="Код", index=False, freeze_panes=(1, 0)
            )

        # Исключаем из filtered_df все строки, где kod присутствует в df_to_exclude
        initial_count = len(filtered_df)
        filtered_df = filtered_df[~filtered_df["kod"].isin(df_to_exclude["kod"])]
        removed_count = initial_count - len(filtered_df)

        # Печатаем количество удаленных строк
        print(f"Удалено строк из filtered_df: {removed_count}")

    except Exception as e:
        print(f"Ошибка при обработке файла: {e}")

    return filtered_df


def load_brands_and_text():
    """
    Функция для поиска файла 'Непереоценивать.xlsx' по указанным путям.
    Если файл найден, извлекает данные с листа 'Бренды' и 'Текст', если они существуют.

    Returns:
    tuple: Два списка - brands и text. Если файл не найден или лист отсутствует, возвращает пустые списки для отсутствующих листов.
    """
    # Пути к файлу
    paths = [
        r"D:\NAS\заказы\Непереоценивать.xlsx",
        r"\\26.218.196.12\заказы\Непереоценивать.xlsx",
    ]

    # Попытка найти файл по одному из путей
    file_path = None
    for path in paths:
        if os.path.exists(path):
            file_path = path
            break

    # Если файл не найден, возвращаем пустые списки
    if file_path is None:
        print("Файл 'Непереоценивать.xlsx' не найден по указанным путям.")
        return [], []

    try:
        # Проверяем, какие листы доступны в файле
        xls = pd.ExcelFile(file_path)
        sheet_names = xls.sheet_names

        # Загружаем данные с листа 'Бренды' и преобразуем первую колонку в список
        if "Бренды" in sheet_names:
            brands_df = pd.read_excel(xls, sheet_name="Бренды", header=None)
            brands = brands_df[0].dropna().astype(str).str.strip().tolist()
        else:
            print("Лист 'Бренды' не найден.")
            brands = []

        # Загружаем данные с листа 'Текст' и преобразуем первую колонку в список
        if "Текст" in sheet_names:
            text_df = pd.read_excel(xls, sheet_name="Текст", header=None)
            text = text_df[0].dropna().astype(str).str.strip().tolist()
        else:
            print("Лист 'Текст' не найден.")
            text = []

        return brands, text

    except Exception as e:
        print(f"Ошибка при обработке файла: {e}")
        return [], []


# Проверка работы функции при запуске скрипта
if __name__ == "__main__":
    from ugkorea.db.database import get_db_engine

    engine = get_db_engine()
    (
        filtered_df,
        priceendmonth_filtered,
        salespivot_filtered,
        stockendmonth_filtered,
        suppliespivot_filtered,
        postuplenija_filtered,
    ) = load_forreprice_data(engine)

    # Вывод для проверки
    print("Filtered DF:")
    print(filtered_df.head())
    print("Price End Month Filtered:")
    print(priceendmonth_filtered.head())
