from db.database import get_db_engine
import pandas as pd
from datetime import datetime, timedelta
import os
from sqlalchemy import create_engine, text
from sqlalchemy.exc import SQLAlchemyError
import re
from transliterate import translit
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Задаем период в месяцах (по умолчанию 12 месяцев)
months = 12
weeks = 1

# Определите путь к выходному файлу
output_file = r'\\26.218.196.12\заказы\Евгений\УпущенныйСпрос.xlsx'

engine = get_db_engine()

def fetch_table_data(engine, table_name):
    try:
        query = text(f"SELECT * FROM {table_name}")
        with engine.connect() as connection:
            result = pd.read_sql(query, connection)
        return result
    except SQLAlchemyError as e:
        print(f"Ошибка при выполнении запроса к таблице {table_name}: {e}")
        return None
    
# Создание новой колонки с информацией об аналогах
def find_analogs(row, df):
    if pd.notna(row['gruppa_analogov']):
        analogs = df[(df['gruppa_analogov'] == row['gruppa_analogov']) & (df['osnsklad'] > 0)]
        if not analogs.empty:
            return ', '.join(analogs['kod'].astype(str) + ' - ' + analogs['proizvoditel'].astype(str) + ' - ' + analogs['osnsklad'].astype(str))
    return None    

def analyze_data(data, period_weeks, months=12):
    """
    Анализирует данные за последние period_weeks и сравнивает их с данными за последние months.
    
    Параметры:
    data (DataFrame): DataFrame с данными.
    period_weeks (int): Период в неделях для анализа.
    months (int): Период в месяцах для анализа (по умолчанию 12).
    
    Возвращает:
    DataFrame: Результирующая таблица с данными.
    """
    # Рассчитываем дату начала периода
    end_date = datetime.now()
    start_date_months = end_date - timedelta(days=months*30)
    start_date_weeks = end_date - timedelta(weeks=period_weeks)

    # Преобразуем столбец с датами в формат datetime
    data['period'] = pd.to_datetime(data['period'], format='%d.%m.%Y')

    # Фильтруем данные по заданному периоду в неделях
    filtered_data_weeks = data[(data['period'] >= start_date_weeks) & (data['period'] <= end_date)]

    # Получаем уникальные коды за период в недели
    unique_codes_weeks = filtered_data_weeks['kod'].unique()

    # Фильтруем данные по заданному периоду в месяцах
    filtered_data_months = data[(data['period'] >= start_date_months) & (data['period'] <= end_date)]

    # Фильтруем данные по уникальным кодам за период в недели
    filtered_data_months = filtered_data_months[filtered_data_months['kod'].isin(unique_codes_weeks)]

    # Группируем данные по колонке 'kod' для периода в месяцах
    grouped_data_months = filtered_data_months.groupby('kod').agg({
        'period': ['count', lambda x: ', '.join(sorted(x.dt.strftime('%d.%m.%Y')))]
    }).reset_index()

    # Переименовываем колонки для читаемости
    grouped_data_months.columns = ['kod', 'Запросы (месяцы)', 'Даты запросов (месяцы)']

    # Группируем данные по колонке 'kod' для периода в неделях
    grouped_data_weeks = filtered_data_weeks.groupby('kod').agg({
        'period': ['count', lambda x: ', '.join(sorted(x.dt.strftime('%d.%m.%Y')))]
    }).reset_index()

    # Переименовываем колонки для читаемости
    grouped_data_weeks.columns = ['kod', 'Запросы (недели)', 'Даты запросов (недели)']

    # Объединяем результаты по 'kod'
    result = pd.merge(grouped_data_months, grouped_data_weeks, on='kod', how='inner')

    # Сортируем по количеству запросов в месяцах и по коду
    result = result.sort_values(by=['Запросы (месяцы)', 'kod'], ascending=[False, True])

    return result

# Функция для удаления знаков препинания и перевода кириллицы в латиницу
def process_article(article):
    if pd.isna(article):
        return article
    # Удаляем знаки препинания
    article = re.sub(r'[^\w\s]', '', article)
    # Переводим кириллицу в латиницу
    article = translit(article, 'ru', reversed=True)
    return article

# Проверка наличия позиций из недельного периода в данных за месячный период
def check_presence_weeks_in_months(row, filtered_data_months):
    article = row['period']
    matches = filtered_data_months[filtered_data_months['period'] == article]
    if not matches.empty:
        dates = ', '.join(matches['period'].dt.strftime('%d.%m.%Y'))
        return pd.Series([len(matches), dates])
    return pd.Series([0, ''])


# Функция для настройки ширины столбцов и переноса текста
def adjust_worksheet(worksheet):
    # Установка переноса текста
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
    
    # Установка удобной ширины столбцов
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter # Получаем букву колонки
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width


# Получаем данные из таблиц nomenklaturaold, priceold, stockold и groupanalogiold
if engine:
    nomenklatura_df = fetch_table_data(engine, 'nomenklaturaold')
    price_df = fetch_table_data(engine, 'priceold')
    stock_df = fetch_table_data(engine, 'stockold')
    groupanalogi_df = fetch_table_data(engine, 'groupanalogiold')

    # Объединяем данные по колонке 'код'
    if nomenklatura_df is not None and price_df is not None and stock_df is not None and groupanalogi_df is not None:
        # Убираем дубликаты по колонке 'код'
        nomenklatura_df = nomenklatura_df.drop_duplicates(subset=['kod'])
        price_df = price_df.drop_duplicates(subset=['kod'])
        stock_df = stock_df.drop_duplicates(subset=['kod'])
        groupanalogi_df = groupanalogi_df.drop_duplicates(subset=['kod_1s', 'gruppa_analogov'])
        
        # Приведение типов данных к строке для гарантии совпадения
        nomenklatura_df['kod'] = nomenklatura_df['kod'].astype(str).str.strip()
        price_df['kod'] = price_df['kod'].astype(str).str.strip()
        stock_df['kod'] = stock_df['kod'].astype(str).str.strip()
        groupanalogi_df['kod_1s'] = groupanalogi_df['kod_1s'].astype(str).str.strip()
        
        # Объединяем таблицы
        merged_df = nomenklatura_df.merge(price_df[['kod', 'tsenazakup','tsenarozn']], on='kod', how='left')
        merged_df = merged_df.merge(stock_df[['kod', 'osnsklad']], on='kod', how='left')

        # Объединяем с таблицей groupanalogiold по колонкам 'код' и 'Код 1С'
        merged_df = merged_df.merge(groupanalogi_df[['kod_1s', 'gruppa_analogov']], left_on='kod', right_on='kod_1s', how='left')

        # Убираем колонку 'Код 1С'
        merged_df.drop(columns=['kod_1s'], inplace=True)

        actual_price = merged_df
        
        # Получаем данные из таблицы upuschennyjspros
        data = fetch_table_data(engine, 'upuschennyjspros')


actual_price['kod'] = actual_price['kod'].str.strip().astype(str)
# Применяем функцию и создаем новую колонку 'Аналоги'
actual_price['Аналоги'] = actual_price.apply(lambda row: find_analogs(row, actual_price), axis=1)



# Рассчитываем дату начала периода
end_date = datetime.now()
start_date = end_date - timedelta(days=months*30)

# Преобразуем столбец с датами в формат datetime
data['period'] = pd.to_datetime(data['period'], format='%d.%m.%Y')

# Фильтруем данные по заданному периоду
filtered_data = data[(data['period'] >= start_date) & (data['period'] <= end_date)]

# Группируем данные по колонке 'Код'
grouped_data = filtered_data.groupby('kod').agg({
    'period': ['count', lambda x: ', '.join(x.dt.strftime('%d.%m.%Y'))]
}).reset_index()

# Переименовываем колонки для читаемости
grouped_data.columns = ['kod', 'Запросы', 'Даты запросов']

# Сортируем по количеству запросов и по коду
grouped_data = grouped_data.sort_values(by=['Запросы', 'kod'], ascending=[False, True])

# Фильтрация данных actual_price для исключения оснсклад > 0
filtered_actual_price = actual_price[actual_price['osnsklad'] == 0].copy()

# Очистка значений в столбцах 'Код' и 'код'
grouped_data['kod'] = grouped_data['kod'].str.strip()
filtered_actual_price['kod'] = filtered_actual_price['kod'].str.strip()

# Объединение таблиц по полю 'Код' и 'код' (inner join)
merged_data = pd.merge(grouped_data, filtered_actual_price, left_on='kod', right_on='kod', how='inner')


union_spros = merged_data


spros_weeks = analyze_data(data, period_weeks=weeks, months=months)

# Очистка данных
spros_weeks['kod'] = spros_weeks['kod'].str.strip().astype(str)

# Объединение таблиц по полю 'Код' и 'код' (left join)
merged_data = pd.merge(spros_weeks, actual_price, on='kod', how='left')

# Удаление лишних колонок после объединения
weeks_spros = merged_data.drop(columns=['pometkaudalenija', 'osnsklad', 'gruppa_analogov'])

# Применяем функцию к колонке АртикулДляПоиска, если Код отсутствует
data.loc[data['kod'].isna(), 'artikuldljapoiska'] = data.loc[data['kod'].isna(), 'artikuldljapoiska'].apply(process_article)

# Преобразуем колонку 'артикул' в actual_price с удалением знаков препинания
actual_price_cleaned = actual_price['artikul'].apply(lambda x: re.sub(r'[^\w\s]', '', str(x)))

# Преобразуем колонку 'АртикулДляПоиска' в data с удалением знаков препинания
data['artikuldljapoiska_cleaned'] = data['artikuldljapoiska'].apply(lambda x: re.sub(r'[^\w\s]', '', str(x)) if pd.notna(x) else x)

# Создаем множество артикулов для быстрого поиска
articul_set = set(actual_price_cleaned)

# Применяем проверку и фильтруем строки, оставляя только те, у которых Код отсутствует
data['exists_in_actual_price'] = data['artikuldljapoiska_cleaned'].apply(lambda x: x in articul_set if pd.notna(x) else False)
data_filtered = data[data['kod'].isna() & ~data['exists_in_actual_price']].drop(columns=['exists_in_actual_price', 'kod'])

# Оставляем только необходимые колонки
data_filtered = data_filtered[['period', 'kolichestvo', 'avtor', 'artikuldljapoiska_cleaned']]

# Переименовываем колонку 'АртикулДляПоиска_cleaned' в 'Артикул'
data_filtered = data_filtered.rename(columns={'artikuldljapoiska_cleaned': 'artikul'})

# Преобразуем колонку 'Период' в datetime
data_filtered['period'] = pd.to_datetime(data_filtered['period'], format='%d.%m.%Y')

# Определяем конечную дату и начальную дату
end_date = data_filtered['period'].max()
start_date = end_date - pd.DateOffset(months=months)

# Фильтруем данные за последние n месяцев
filtered_data = data_filtered[(data_filtered['period'] >= start_date) & (data_filtered['period'] <= end_date)]

# Группируем данные по 'Артикул', подсчитываем количество запросов и объединяем даты запросов
grouped_data = filtered_data.groupby('artikul').agg(
    Запросы=('artikul', 'count'),
    Даты_запросов=('period', lambda x: ', '.join(x.dt.strftime('%d.%m.%Y')))
).reset_index()

# Сортируем по количеству запросов в убывающем порядке
grouped_data = grouped_data.sort_values(by='Запросы', ascending=False)


# Определяем конечную дату и начальную дату для недельного периода
end_date_weeks = data_filtered['period'].max()
start_date_weeks = end_date_weeks - pd.DateOffset(weeks=weeks)

# Фильтруем данные за последние n недель
filtered_data_weeks = data_filtered[(data_filtered['period'] >= start_date_weeks) & (data_filtered['period'] <= end_date_weeks)]

# Фильтруем данные за последние n месяцев
start_date_months = end_date_weeks - pd.DateOffset(months=months)
filtered_data_months = data_filtered[(data_filtered['period'] >= start_date_months) & (data_filtered['period'] <= end_date_weeks)]


# Применяем проверку и добавляем результаты в новый DataFrame
result = filtered_data_weeks.copy()
result[['Количество за месяцы', 'Даты запросов за месяцы']] = result.apply(lambda row: check_presence_weeks_in_months(row, filtered_data_months), axis=1)

# Оставляем только необходимые колонки
result = result[['period', 'kolichestvo', 'avtor', 'artikul', 'Количество за месяцы', 'Даты запросов за месяцы']]


assert 'union_spros' in globals(), "union_spros не найден"
assert 'weeks_spros' in globals(), "weeks_spros не найден"
assert 'grouped_data' in globals(), "grouped_data не найден"
assert 'result' in globals(), "result не найден"
assert 'actual_price' in globals(), "actual_price не найден"


with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    union_spros.to_excel(writer, sheet_name='общий', index=False)
    weeks_spros.to_excel(writer, sheet_name='запериод', index=False)
    grouped_data.to_excel(writer, sheet_name='НПобщий', index=False)
    result.to_excel(writer, sheet_name='НПпериод', index=False)
    actual_price.to_excel(writer, sheet_name='ОбщийПрайссгруппами', index=False)

# Загрузка созданного файла для внесения изменений
wb = load_workbook(output_file)


# Применение функции ко всем листам
for sheet_name in wb.sheetnames:
    adjust_worksheet(wb[sheet_name])

# Сохранение изменений
wb.save(output_file)

print(f"Файл успешно создан и сохранен как '{output_file}'")